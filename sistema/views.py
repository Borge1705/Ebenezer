# =========================
# DJANGO CORE
# =========================
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Sum, DecimalField
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from django.conf import settings

# =========================
# MODELOS
# =========================
from .models import (
    Telefono,
    Reparacion,
    Cliente,
    Venta,
    DetalleVenta,
    Accesorio,
)

# =========================
# FORMULARIOS
# =========================
from .forms import *

# =========================
# EXPORT EXCEL
# =========================
from openpyxl import Workbook

# =========================
# PDF REPORTS (REPORTLAB)
# =========================
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    Paragraph,
    Spacer,
    TableStyle,
    Image,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter

# =========================
# GRÁFICOS
# =========================
import matplotlib.pyplot as plt

# =========================
# UTILIDADES
# =========================
from io import BytesIO
import json
import os
from datetime import date
from django.utils.timezone import now
C_PRIMARY = colors.HexColor("#0f3460")
C_SECONDARY = colors.HexColor("#1565C0")
C_ACCENT = colors.HexColor("#1E88E5")
C_BG = colors.HexColor("#F5F7FA")
C_GRID = colors.HexColor("#CBD5E1")

LOGO = os.path.join(
    settings.BASE_DIR,
    "static",
    "sistema",
    "img",
    "logo.jpeg"
)


def es_admin_o_vendedor(user):

    return (
        user.is_superuser
        or user.groups.filter(
            name="Vendedores"
        ).exists()
    )

@login_required
def dashboard_vendedor(request):

    return render(
        request,
        "sistema/dashboard_vendedor.html"
    )

def es_tecnico(user):
    return user.groups.filter(
        name='Tecnico'
    ).exists()

def obtener_estadisticas():


    total_money = Venta.objects.aggregate(
        total=Coalesce(
            Sum('total'),
            0,
            output_field=DecimalField()
        )
    )['total']



    ventas_hoy = Venta.objects.filter(
        fecha_venta=date.today()
    ).aggregate(
        total=Coalesce(
            Sum('total'),
            0,
            output_field=DecimalField()
        )
    )['total']



    return {


        # GENERALES

        'total_phones':
            Telefono.objects.count(),


        'total_accessories':
            Accesorio.objects.count(),


        'total_clients':
            Cliente.objects.count(),


        'total_money':
            total_money,



        # REPARACIONES

        'active_repairs':
            Reparacion.objects.exclude(
                estado='Completado'
            ).count(),


        'completed_repairs':
            Reparacion.objects.filter(
                estado='Completado'
            ).count(),


        'delivered_repairs':
            Reparacion.objects.filter(
                estado='Entregado'
            ).count(),



        'pending_repairs':
            Reparacion.objects.filter(
                estado='Pendiente'
            ).count(),


        'process_repairs':
            Reparacion.objects.filter(
                estado='En proceso'
            ).count(),



        # STOCK

        'out_of_stock_phones':
            Telefono.objects.filter(
                estado='Agotado'
            ).count(),


        'low_stock_phones':
            Telefono.objects.filter(
                stock__lte=3,
                stock__gt=0
            ).count(),



        'low_stock_accessories':
            Accesorio.objects.filter(
                stock__lte=3,
                stock__gt=0
            ).count(),



        # VENTAS

        'sales_today':
            ventas_hoy,



        # LISTAS

        'recent_sales':
            Venta.objects.order_by(
                '-fecha_venta'
            )[:5],


        'recent_repairs':
            Reparacion.objects.order_by(
                '-fecha'
            )[:5],


        'recent_clients':
            Cliente.objects.order_by(
                '-id'
            )[:5],

    }

# LOGIN

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user:

            login(request, user)

            if es_tecnico(user):
                return redirect('dashboard_tecnico')
            

            if user.groups.filter(name='Vendedores').exists():
                return redirect('dashboard_vendedor')

            return redirect('dashboard')

    return render(request, 'sistema/login.html')


# LOGOUT

def logout_view(request):
    logout(request)
    return redirect('login')


# DASHBOARD

@login_required
def dashboard(request):

    context = obtener_estadisticas()


    context.update({

        "meses":[
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio"
        ],


        "ventas":[
            0,
            0,
            0,
            0,
            0,
            context['total_money']
        ]

    })


    return render(
        request,
        'sistema/dashboard.html',
        context
    )


@login_required
def dashboard_tecnico(request):

    context = {

        'reparaciones_pendientes':
        Reparacion.objects.exclude(
            estado='Completado'
        ).count(),

        'reparaciones_completadas':
        Reparacion.objects.filter(
            estado='Completado'
        ).count(),

        'telefonos':
        Telefono.objects.count(),

        'accesorios':
        Accesorio.objects.count(),
    }

    return render(
        request,
        'sistema/dashboard_tecnico.html',
        context
    )

@login_required
def inventario(request):

    # EDITAR TELEFONO
    if request.method == 'POST' and 'edit_id' in request.POST:

        telefono = Telefono.objects.get(id=request.POST['edit_id'])

        telefono.modelo = request.POST['modelo']
        telefono.marca = request.POST['marca']
        telefono.precio = request.POST['precio']
        telefono.stock = int(request.POST['stock'])
        if request.FILES.get('imagen'):
            telefono.imagen = request.FILES.get('imagen')

        telefono.save()

        return redirect('inventario')

    # NUEVO TELEFONO
    elif request.method == 'POST':

        Telefono.objects.create(
            modelo=request.POST['modelo'],
            marca=request.POST['marca'],
            precio=request.POST['precio'],
            stock=int(request.POST['stock']),
            imagen=request.FILES.get('imagen')
        )

        return redirect('inventario')

    phones = Telefono.objects.all()

    context = {
        'phones': phones,
        **obtener_estadisticas()
    }

    return render(request, 'sistema/inventario.html', context)

@login_required
def clientes(request):

    # CREAR O EDITAR
    if request.method == 'POST':

        # EDITAR
        if request.POST.get('edit_cliente_id'):

            cliente = Cliente.objects.get(
                id=request.POST.get('edit_cliente_id')
            )

            cliente.nombre = request.POST['nombre']
            cliente.telefono = request.POST['telefono']
            cliente.email = request.POST['email']

            cliente.save()

        # NUEVO
        else:

            Cliente.objects.create(

                nombre=request.POST['nombre'],
                telefono=request.POST['telefono'],
                email=request.POST['email']

            )

        return redirect('clientes')

    clientes = Cliente.objects.all()

    context = {
        'clientes': clientes,
        **obtener_estadisticas()
    }

    return render(request, 'sistema/clientes.html', context)


@login_required
@user_passes_test(es_admin_o_vendedor)
def generar_orden(request):

    if request.method == "POST":


        Reparacion.objects.create(

            cliente_id=request.POST['cliente'],

            marca=request.POST['marca'],

            modelo=request.POST['modelo'],

            serie=request.POST['serie'],

            accesorios_recibidos=request.POST['accesorios'],


            defecto_cliente=request.POST['defecto'],


            notas=request.POST['notas'],


            informacion_taller=request.POST['taller'],


            estado="Pendiente"

        )


        return redirect(
            'generar_orden'
        )



    ordenes = Reparacion.objects.order_by('-fecha')


    return render(
        request,
        'sistema/generar_orden.html',
        {
            'clientes': Cliente.objects.all(),
            'telefonos': Telefono.objects.all(),
            'ordenes': ordenes
        }
    )

@login_required
def reparaciones(request):

    data = Reparacion.objects.all()


    if request.method == 'POST':

        form = ReparacionForm(request.POST)


        if form.is_valid():

            reparacion = form.save(commit=False)

            # Cuando se crea una reparación queda pendiente
            reparacion.estado = "Pendiente"

            reparacion.save()


            return redirect('reparaciones')


    else:

        form = ReparacionForm()



    return render(
        request,
        'sistema/reparaciones.html',
        {
            'reparaciones': data,
            'form': form
        }
    )


#PARTE DEL TECNICO DE LAS REPARACIONES
@login_required
def ordenes_tecnico(request):

    ordenes = Reparacion.objects.all()


    return render(
        request,
        'sistema/ordenes_tecnico.html',
        {
            'ordenes': ordenes
        }
    )



@login_required
def detalle_orden(request, id):

    reparacion = get_object_or_404(
        Reparacion,
        id=id
    )


    if request.method == "POST":


        reparacion.cotizacion_servicio = request.POST.get(
            'cotizacion_servicio'
        )


        reparacion.precio_cotizacion = request.POST.get(
            'precio_cotizacion'
        )


        reparacion.save()


        return redirect(
            'detalle_orden',
            id=id
        )



    return render(
        request,
        'sistema/detalle_orden.html',
        {
            'reparacion': reparacion
        }
    )


@login_required
def editar_orden_tecnico(request, id):

    reparacion = get_object_or_404(
        Reparacion,
        id=id
    )


    if request.method == "POST":

        reparacion.caracteristicas = request.POST['caracteristicas']

        reparacion.razon_visita = request.POST['razon_visita']

        reparacion.observaciones = request.POST['observaciones']

        reparacion.diagnostico = request.POST['diagnostico']

        reparacion.proceso_realizado = request.POST['proceso_realizado']

        reparacion.estado = request.POST['estado']


        reparacion.tecnico = request.user


        reparacion.save()


        return redirect(
            'ordenes_tecnico'
        )


    return render(
        request,
        'sistema/editar_orden_tecnico.html',
        {
            'reparacion': reparacion
        }
    )
from django.shortcuts import render, redirect, get_object_or_404
from .models import *


# ================= INVENTARIO =================

def nuevo_telefono(request):
    return render(request, 'sistema/nuevo_telefono.html')


def editar_telefono(request, id):
    telefono = get_object_or_404(Telefono, id=id)

    context = {
        'telefono': telefono
    }

    return render(request, 'sistema/editar_telefono.html', context)


def eliminar_telefono(request, id):
    telefono = get_object_or_404(Telefono, id=id)
    telefono.delete()

    return redirect('inventario')


# ================= CLIENTES =================

def nuevo_cliente(request):
    return render(request, 'sistema/nuevo_cliente.html')


def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    context = {
        'cliente': cliente
    }

    return render(request, 'sistema/editar_cliente.html', context)

def eliminar_cliente(request, id):

    cliente = Cliente.objects.get(id=id)

    cliente.delete()

    return redirect('clientes')




# ================= REPARACIONES =================

def nueva_reparacion(request):
    return render(request, 'sistema/nueva_reparacion.html')


def editar_reparacion(request, id):
    reparacion = get_object_or_404(Reparacion, id=id)

    context = {
        'reparacion': reparacion
    }

    return render(request, 'sistema/editar_reparacion.html', context)


def eliminar_reparacion(request, id):
    reparacion = get_object_or_404(Reparacion, id=id)
    reparacion.delete()

    return redirect('reparaciones')



# =========================
# VENTAS
# =========================

@login_required
@user_passes_test(es_admin_o_vendedor)
def ventas(request):

    clientes = Cliente.objects.all()

    telefonos = Telefono.objects.filter(
        stock__gt=0
    )

    accesorios = Accesorio.objects.filter(
        stock__gt=0
    )


    if request.method == "POST":

        cliente = request.POST["cliente"]

        productos = request.POST.getlist(
            "productos[]"
        )

        cantidades = request.POST.getlist(
            "cantidades[]"
        )


        with transaction.atomic():

            venta = Venta.objects.create(
                cliente_id=cliente,
                total=0,
                usuario=request.user
            )


            total = 0


            for producto_id, cantidad in zip(productos,cantidades):

                cantidad = int(cantidad)


                if producto_id.startswith("tel_"):


                    telefono_id = producto_id.replace(
                        "tel_",
                        ""
                    )


                    producto = Telefono.objects.select_for_update().get(
                        id=telefono_id
                    )


                    if producto.stock < cantidad:
                        venta.delete()
                        return redirect("ventas")


                    subtotal = producto.precio * cantidad


                    producto.stock -= cantidad
                    producto.save()



                    DetalleVenta.objects.create(

                        venta=venta,

                        producto=f"{producto.marca} {producto.modelo}",

                        tipo_producto="Telefono",

                        cantidad=cantidad,

                        precio_unitario=producto.precio,

                        subtotal=subtotal
                    )



                else:


                    accesorio_id = producto_id.replace(
                        "acc_",
                        ""
                    )


                    producto = Accesorio.objects.select_for_update().get(
                        id=accesorio_id
                    )



                    if producto.stock < cantidad:
                        venta.delete()
                        return redirect("ventas")



                    subtotal = producto.precio * cantidad



                    producto.stock -= cantidad
                    producto.save()



                    DetalleVenta.objects.create(

                        venta=venta,

                        producto=producto.nombre,

                        tipo_producto="Accesorio",

                        cantidad=cantidad,

                        precio_unitario=producto.precio,

                        subtotal=subtotal
                    )



                total += subtotal



            venta.total = total

            venta.save()



        return redirect("ventas")



    ventas = Venta.objects.prefetch_related(
        "detalles"
    ).order_by("-id")



    return render(
        request,
        "sistema/ventas.html",
        {

        "ventas":ventas,

        "clientes":clientes,

        "telefonos":telefonos,

        "accesorios":accesorios,

        **obtener_estadisticas()

        }
    )


# =========================
# ELIMINAR VENTA
# =========================

@login_required
def eliminar_venta(request, id):

    venta = get_object_or_404(
        Venta,
        id=id
    )

    venta.delete()

    return redirect('ventas')

# =========================
# ACCESORIOS
# =========================

from .models import Accesorio

@login_required
def accesorios(request):

    # EDITAR ACCESORIO
    if request.method == 'POST' and request.POST.get('edit_id'):

        accesorio = Accesorio.objects.get(
            id=request.POST.get('edit_id')
        )

        accesorio.nombre = request.POST['nombre']
        accesorio.tipo = request.POST['tipo']
        accesorio.precio = request.POST['precio']
        accesorio.stock = request.POST['stock']

        if request.FILES.get('imagen'):
            accesorio.imagen = request.FILES.get('imagen')


        accesorio.save()

        return redirect('accesorios')

    # NUEVO ACCESORIO
    elif request.method == 'POST':

        Accesorio.objects.create(
            nombre=request.POST['nombre'],
            tipo=request.POST['tipo'],
            precio=request.POST['precio'],
            stock=request.POST['stock'],
            imagen=request.FILES.get('imagen')

        )

        return redirect('accesorios')

    accesorios = Accesorio.objects.all()

    context = {
        'accesorios': accesorios,
        **obtener_estadisticas()
    }

    return render(request, 'sistema/accesorios.html', context)

# ELIMINAR ACCESORIO

def eliminar_accesorio(request, id):

    accesorio = Accesorio.objects.get(id=id)

    accesorio.delete()

    return redirect('accesorios')


####REPORTES DE VENTAS, INVENTARIO Y REPARACIONES

@login_required
def reportes(request):

    return render(
        request,
        'sistema/reportes.html'
    )


@login_required
def reporte_ventas_excel(request):

    ventas = Venta.objects.prefetch_related('detalles').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append([
        "ID Venta", "Cliente", "Producto", "Tipo",
        "Cantidad", "Precio", "Subtotal", "Fecha"
    ])

    for venta in ventas:
        for detalle in venta.detalles.all():
            ws.append([
                venta.id,
                venta.cliente.nombre,
                detalle.producto,
                detalle.tipo_producto,
                detalle.cantidad,
                detalle.precio_unitario,
                detalle.subtotal,
                venta.fecha_venta
            ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas.xlsx"'

    wb.save(response)
    return response


@login_required
def reporte_inventario_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(["Producto", "Tipo", "Precio", "Stock", "Estado"])

    for t in Telefono.objects.all():
        ws.append([
            f"{t.marca} {t.modelo}",
            "Telefono",
            t.precio,
            t.stock,
            t.estado
        ])

    for a in Accesorio.objects.all():
        estado = "Agotado" if a.stock == 0 else "Disponible"
        ws.append([
            a.nombre,
            "Accesorio",
            a.precio,
            a.stock,
            estado
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Inventario.xlsx"'

    wb.save(response)
    return response


@login_required
def reporte_inventario_pdf(request):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE INVENTARIO - EBENEZER", estilos["Title"]))
    elementos.append(Spacer(1, 10))

    datos = [["Producto", "Tipo", "Stock"]]

    telefonos = Telefono.objects.all()
    accesorios = Accesorio.objects.all()

    labels = []
    valores = []

    for t in telefonos:
        datos.append([t.modelo, "Telefono", t.stock])
        labels.append(t.modelo)
        valores.append(t.stock)

    for a in accesorios:
        datos.append([a.nombre, "Accesorio", a.stock])
        labels.append(a.nombre)
        valores.append(a.stock)

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_PRIMARY),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, C_GRID),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, C_BG])
    ]))

    elementos.append(tabla)

    # gráfico moderno
    plt.figure(figsize=(6,3))
    plt.bar(labels[:8], valores[:8], color="#1565C0")
    plt.xticks(rotation=45)
    plt.title("Stock disponible")

    grafica = BytesIO()
    plt.tight_layout()
    plt.savefig(grafica, format="png")
    plt.close()

    grafica.seek(0)
    elementos.append(Image(grafica, width=350, height=180))

    doc.build(elementos)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Inventario_Ebenezer.pdf"'

    return response

# =========================
# REPORTE VENTAS PDF
# =========================

@login_required
def reporte_ventas_pdf(request):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE VENTAS - EBENEZER", estilos["Title"]))
    elementos.append(Spacer(1, 10))

    ventas = Venta.objects.prefetch_related("detalles")

    for venta in ventas:

        elementos.append(
            Paragraph(f"Factura #{venta.id} - Cliente: {venta.cliente.nombre}", estilos["Heading2"])
        )

        datos = [["Producto", "Cantidad", "Precio", "Subtotal"]]

        for d in venta.detalles.all():
            datos.append([
                d.producto,
                d.cantidad,
                f"${d.precio_unitario}",
                f"${d.subtotal}"
            ])

        tabla = Table(datos)

        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), C_PRIMARY),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 0.5, C_GRID),
            ("ALIGN", (0,0), (-1,-1), "CENTER")
        ]))

        elementos.append(tabla)
        elementos.append(Spacer(1, 15))

    doc.build(elementos)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Ventas_Ebenezer.pdf"'

    return response


from django.utils import timezone
from openpyxl import Workbook
from django.http import HttpResponse


@login_required
def reporte_reparaciones_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Reparaciones"

    ws.append([
        "Orden", "Cliente", "Equipo",
        "Estado", "Tecnico", "Fecha"
    ])

    reparaciones = Reparacion.objects.all()

    for r in reparaciones:

        if r.fecha:
            fecha = r.fecha
            if timezone.is_aware(fecha):
                fecha = timezone.make_naive(fecha)
        else:
            fecha = ""

        ws.append([
            r.id,
            r.cliente.nombre,
            r.modelo,
            r.estado,
            r.tecnico.username if r.tecnico else "",
            fecha
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Reparaciones.xlsx"'

    wb.save(response)
    return response


@login_required
def reporte_reparaciones_pdf(request):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)

    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE REPARACIONES - EBENEZER", estilos["Title"]))
    elementos.append(Spacer(1, 10))

    datos = [["Orden", "Cliente", "Equipo", "Estado", "Cotización"]]

    for r in Reparacion.objects.all():

        datos.append([
            r.id,
            r.cliente.nombre,
            r.modelo,
            r.estado,
            f"${getattr(r,'cotizacion',0)}"
        ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_PRIMARY),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, C_GRID),
        ("ALIGN", (0,0), (-1,-1), "CENTER")
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Reparaciones_Ebenezer.pdf"'

    return response

# ================================
# REPORTE GENERAL PDF
# ================================

@login_required
def reporte_general_pdf(request):


    buffer = BytesIO()


    doc = SimpleDocTemplate(
        buffer,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )


    elementos=[]


    estilos=getSampleStyleSheet()



    # ==========================
    # LOGO
    # ==========================


    if os.path.exists(LOGO):

        img = Image(
            LOGO,
            width=80,
            height=80
        )

        elementos.append(img)



    elementos.append(
        Paragraph(
            "Reporte General del Sistema - EBENEZER",
            estilos["Title"]
        )
    )



    elementos.append(
        Paragraph(
            f"Fecha: {timezone.now().date()}",
            estilos["Normal"]
        )
    )


    elementos.append(
        Spacer(1,20)
    )



    # ==========================
    # DATOS GENERALES
    # ==========================


    total_ventas = Venta.objects.count()


    ingresos = sum(
        v.total for v in Venta.objects.all()
    )



    productos = sum(
        d.cantidad
        for d in DetalleVenta.objects.all()
    )



    bajo_stock = (
        Telefono.objects.filter(stock__lte=3).count()
        +
        Accesorio.objects.filter(stock__lte=3).count()
    )



    reparaciones = Reparacion.objects.exclude(
        estado="Finalizado"
    ).count()



    datos=[

        ["Indicador","Cantidad"],

        [
            "Ventas realizadas",
            total_ventas
        ],

        [
            "Ingresos",
            "$"+str(ingresos)
        ],

        [
            "Productos vendidos",
            productos
        ],

        [
            "Stock bajo",
            bajo_stock
        ],

        [
            "Reparaciones activas",
            reparaciones
        ]

    ]



    tabla=Table(datos)



    tabla.setStyle(
        TableStyle([

            ("BACKGROUND",
             (0,0),
             (-1,0),
             colors.HexColor("#0f3460")),

            ("TEXTCOLOR",
             (0,0),
             (-1,0),
             colors.white),

            ("GRID",
             (0,0),
             (-1,-1),
             1,
             colors.grey)

        ])
    )



    elementos.append(tabla)



    elementos.append(
        Spacer(1,25)
    )



    # ==========================
    # GRAFICA CAJA
    # ==========================


    ventas = Venta.objects.all()


    meses={}
    

    for v in ventas:

        mes=v.fecha_venta.strftime("%m")

        meses[mes]=meses.get(
            mes,
            0
        ) + float(v.total)



    if meses:


        plt.figure(figsize=(5,3))


        plt.bar(
            meses.keys(),
            meses.values()
        )


        plt.title(
            "Ingresos por mes"
        )


        plt.xlabel(
            "Mes"
        )


        plt.ylabel(
            "Ingresos"
        )


        grafica1=BytesIO()


        plt.savefig(
            grafica1,
            format="png",
            bbox_inches="tight"
        )


        plt.close()



        grafica1.seek(0)



        elementos.append(
            Image(
                grafica1,
                width=300,
                height=180
            )
        )



    elementos.append(
        Spacer(1,25)
    )



    # ==========================
    # GRAFICA REPARACIONES
    # ==========================


    estados={}



    for r in Reparacion.objects.all():


        estados[r.estado]=(
            estados.get(
                r.estado,
                0
            )+1
        )



    if estados:


        plt.figure(figsize=(4,4))


        plt.pie(
            estados.values(),
            labels=estados.keys(),
            autopct="%1.1f%%"
        )


        plt.title(
            "Estado de reparaciones"
        )



        grafica2=BytesIO()


        plt.savefig(
            grafica2,
            format="png",
            bbox_inches="tight"
        )


        plt.close()


        grafica2.seek(0)


        elementos.append(
            Image(
                grafica2,
                width=250,
                height=250
            )
        )



    doc.build(elementos)



    pdf=buffer.getvalue()


    buffer.close()



    response=HttpResponse(
        pdf,
        content_type="application/pdf"
    )


    response["Content-Disposition"]=(
        'attachment; filename="Reporte_General_Ebenezer.pdf"'
    )


    return response

def encabezado_pdf(titulo):

    elementos = []

    estilos = getSampleStyleSheet()


    if os.path.exists(LOGO):

        img = Image(
            LOGO,
            width=80,
            height=80
        )

        elementos.append(img)



    elementos.append(
        Paragraph(
            f"{titulo} - EBENEZER",
            estilos["Title"]
        )
    )


    elementos.append(
        Paragraph(
            f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            estilos["Normal"]
        )
    )


    elementos.append(
        Spacer(1,20)
    )


    return elementos